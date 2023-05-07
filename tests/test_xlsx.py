import os.path
from openpyxl import load_workbook
from property import RESOURCES_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    xlsx_path = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
    assert sheet.cell(row=3, column=3).value == 'Hashimoto'
    assert sheet.cell(row=3, column=5).value == 'Great Britain'
    assert sheet.cell(row=3, column=8).value == 1582


